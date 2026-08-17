"""
The seven source feeds.

Batch  (cloud files): EHR extract, claims extract, pharmacy inventory, bed capacity
Batch  (SharePoint) : staff schedules, as .xlsx with realistic human mess
Stream (Kafka)      : patient vitals, prescription issuance

Field names follow the schemas in data-generation-schema-spec.md. Where a field
is our own addition it is noted in that document, not invented here.
"""

import io
import math
import random
import uuid
from collections import defaultdict
from datetime import date, datetime, timedelta, time

from . import config as C
from . import refdata as R


def _iso(dt):
    return dt.strftime("%Y-%m-%dT%H:%M:%SZ") if dt else None


def _dstr(d):
    return d.strftime("%Y%m%d")


def _case(value, source_system):
    """
    Apply the emitting system's enum casing.

    Seven facilities on a mix of EHRs do not agree on casing, so the same
    logical value arrives as `emergency`, `EMERGENCY` and `Emergency`. Silver
    has to fold them; Bronze must preserve them.
    """
    if value is None:
        return None
    style = C.EHR_ENUM_CASE.get(source_system, "lower")
    if style == "upper":
        return str(value).upper()
    if style == "title":
        return str(value).title()
    return str(value).lower()


def _sysdate(dt, source_system):
    """
    Render a timestamp in the emitting system's own format.

    This is why a Bronze layer stores everything as STRING. Four different
    formats arrive, including one day-first (`%d/%m/%Y`) that will parse
    silently and wrongly if a reader assumes month-first.
    """
    if dt is None:
        return None
    fmt = C.EHR_DATE_FORMATS.get(source_system)
    return dt.strftime(fmt) if fmt else dt


class Emitters:
    def __init__(self, sim, dims, batch_sink, stream_sink, injector, rng_seed=C.SEED,
                 emit_from=None):
        self.sim = sim
        self.dims = dims
        self.batch = batch_sink
        self.stream = stream_sink
        self.dx = injector
        self.rng = random.Random(rng_seed + 7)
        self.stats = defaultdict(int)
        # Earliest encounter close date that made it into the EHR feed. Every
        # feed except claims looks back exactly one day, so they are all
        # mutually consistent by construction; claims looks back
        # CODING_LAG_DAYS and would otherwise bill discharges from the warmup
        # period that no EHR extract ever contained -- manufacturing tens of
        # thousands of orphan foreign keys and drowning out the handful the
        # defect injector adds on purpose.
        self.emit_from = emit_from
        # A hire date is a property of the person, not of the extract that
        # reports them. Deriving it from run_date made every hire_date and
        # termination_date in dim_staff slide forward on each weekly refresh,
        # so all ~5k staff looked like they had changed -- which would produce
        # thousands of phantom SCD-2 versions a week and silently wrong tenure.
        # Latch one epoch for the whole run instead.
        self._staff_epoch = emit_from
        # inventory state carried across days
        self._inventory = None
        self._shortages = {}

    def _write(self, path, rows, fieldnames=None):
        """
        Write a fact-feed file, allowing the drop itself to fail.

        Reference dimensions deliberately do NOT go through here. A missing
        dim_facility would orphan every fact row in the batch and bury the
        handful of orphan foreign keys the injector adds on purpose, so the
        referential-integrity signal would be destroyed rather than tested.
        """
        action, payload = self.dx.file_failure(path, rows)
        if action == "missing":
            self.stats["files_missing"] += 1
            return 0
        if action == "truncated":
            self.stats["files_truncated"] += 1
        return self.batch.write_csv(path, payload, fieldnames)

    # =====================================================================
    # Dimensions (weekly full refresh)
    # =====================================================================
    def emit_dimensions(self, run_date):
        d = _dstr(run_date)

        self.batch.write_csv(f"reference/dim_facility/dim_facility_{d}.csv", [
            {
                "facility_id": f.facility_id, "facility_name": f.name,
                "facility_type": f.facility_type, "region": f.region,
                "city": f.city, "state": f.state, "zip": f.zip, "county": f.county,
                "emergency_services": f.has_ed, "licensed_beds": f.licensed_beds,
                "staffed_beds": sum(u.staffed_beds for u in self.dims.units_by_facility.get(f.facility_id, [])),
                "ownership": f.ownership,
            } for f in self.dims.facilities])

        self.batch.write_csv(f"reference/dim_unit/dim_unit_{d}.csv", [
            {
                "unit_id": u.unit_id, "facility_id": u.facility_id,
                "unit_code": u.unit_code, "unit_name": u.unit_name,
                "unit_type": u.unit_type, "building": u.building, "floor": u.floor,
                "licensed_beds": u.licensed_beds, "staffed_beds": u.staffed_beds,
                "blocked_beds": u.blocked_beds,
                "nurse_patient_ratio_target": u.nurse_patient_ratio_target,
                "is_critical_care": u.is_critical_care,
                "is_monitored": u.is_monitored,
            } for u in self.dims.units])

        self.batch.write_csv(f"reference/dim_payer/dim_payer_{d}.csv", self.dims.payers)

        self.batch.write_csv(f"reference/dim_drug/dim_drug_{d}.csv", self.dims.drugs)

        # fixed for the whole run; latch the first extract date if unset
        if self._staff_epoch is None:
            self._staff_epoch = run_date
        hire_base = self._staff_epoch
        staff_rows = []
        for s in self.dims.staff:
            r = dict(s)
            r["hire_date"] = hire_base + timedelta(days=s["hire_date_offset_days"])
            # A snapshot cannot contain someone who has not been hired yet.
            # Staff churn adds hires dated inside the window, so without this
            # filter the first weekly refresh would list future employees.
            if r["hire_date"] > run_date:
                continue
            r["termination_date"] = (hire_base + timedelta(days=s["termination_offset_days"])
                                     if s["termination_offset_days"] is not None else None)
            # Terminated, but not yet as of this snapshot -- still active here.
            if r["termination_date"] and r["termination_date"] > run_date:
                r["termination_date"] = None
                r["is_active"] = True
            else:
                r["is_active"] = not s["is_terminated"]
            for k in ("hire_date_offset_days", "termination_offset_days",
                      "is_terminated", "is_licensed_nurse"):
                r.pop(k, None)
            staff_rows.append(r)
        self.batch.write_csv(f"reference/dim_staff/dim_staff_{d}.csv", staff_rows)

    # =====================================================================
    # Code-set reference feeds
    #
    # The real code sets are embedded in the fact rows -- ICD-10-CM on
    # diagnoses, MS-DRG on claims -- but were never emitted as lookups, so a
    # diagnosis dimension had nothing to build from. Emitted on the same weekly
    # full-refresh cadence as the other reference files.
    # =====================================================================
    CHRONIC_PREFIXES = ("E11", "I10", "I50", "J44", "N18", "I48")

    def emit_code_sets(self, run_date):
        d = _dstr(run_date)

        icd_rows = []
        for code, title, weight, setting, cohort in self.dims.icd10:
            # is_chronic / readmission_risk_level are [OURS]. ICD-10-CM itself
            # carries neither; CMS chronic-condition flags come from the CCW
            # algorithm, which is a separate licensed artefact.
            chronic = code.startswith(self.CHRONIC_PREFIXES)
            risk = "high" if cohort != "OTHER" else ("medium" if chronic else "low")
            icd_rows.append({
                "icd10_code": code,
                "icd10_description": title,
                "icd_version": 10,
                "code_chapter": code[0],
                "diagnosis_category": R.drg_family_for(code),
                "care_setting": setting,
                "hrrp_cohort": cohort,
                "is_chronic": chronic,
                "readmission_risk_level": risk,
                "relative_frequency": weight,
            })
        self.batch.write_csv(f"reference/dim_icd10/dim_icd10_{d}.csv", icd_rows)

        self.batch.write_csv(f"reference/dim_drg/dim_drg_{d}.csv", R.all_drgs())
        self.stats["code_set_rows"] += len(icd_rows) + len(R.all_drgs())

    def _is_alarm(self, loinc, value):
        """
        Would the bedside monitor alarm on this reading?

        True when the value falls in a band that scores 3 on NEWS2 for that
        parameter -- the single-parameter trigger that prompts clinician review.
        Diastolic BP is not a NEWS2 parameter and never alarms.

        NOTE the systolic and SpO2 Scale 2 bands still need confirming against
        the official RCP chart -- see the checklist in the schema spec.
        """
        try:
            v = float(value)
        except (TypeError, ValueError):
            return False
        if loinc == "8867-4":                       # heart rate
            return v <= 40 or v >= 131
        if loinc == "8480-6":                       # systolic BP
            return v <= 90 or v >= 220
        if loinc == "9279-1":                       # respiratory rate
            return v <= 8 or v >= 25
        if loinc == "8310-5":                       # temperature
            return v <= 35.0
        if loinc == "2708-6":                       # SpO2
            return v <= 91
        return False

    # =====================================================================
    # Source 1 — EHR encounters & admissions (daily CSV, cloud files)
    # =====================================================================
    def emit_ehr(self, run_date):
        """Encounters that closed on run_date - 1, i.e. yesterday's activity."""
        day = run_date - timedelta(days=1)
        d = _dstr(run_date)

        encs = self.sim.by_close_date.get(day, [])
        if not encs:
            return

        patients, encounters, admissions, transfers, ed_stays, diagnoses = [], [], [], [], [], []
        outpatient = []
        seen_patients = set()

        for e in encs:
            p = e.patient
            src = C.EHR_SYSTEMS.get(e.facility_id, "MERIDIAN_EHR_CORE")
            mrn = self.dims.mrn_for(p, e.facility_id)
            birth = day - timedelta(days=int(p["_age"] * 365.25) + self.rng.randint(0, 364))

            if p["enterprise_patient_id"] not in seen_patients:
                seen_patients.add(p["enterprise_patient_id"])
                prow = {k: v for k, v in p.items()
                        if k not in ("_age", "mrn_by_facility", "enterprise_patient_id")}
                prow["BirthDate"] = birth
                prow["mrn"] = mrn
                prow["source_facility_id"] = e.facility_id
                prow, dup = self.dx.mutate_row(
                    "ehr.patients", prow, "Id",
                    required_fields=("BirthDate", "Last", "Gender"),
                    coded_fields=("Gender", "Race", "Ethnicity"),
                    numeric_fields=("Income",))
                patients.append(prow)
                if dup:
                    patients.append(dup)

            erow = {
                "Id": e.encounter_id,
                "Start": e.ed_arrival or e.admittime,
                "Stop": e.dischtime or e.ed_departure,
                "Patient": p["Id"],
                "Organization": e.facility_id,
                "Provider": self.rng.choice(self.dims.staff)["staff_id"],
                "Payer": e.payer["payer_id"] if e.payer else None,
                "EncounterClass": e.encounter_class,
                "Code": e.principal_diagnosis[1] if e.principal_diagnosis else None,
                "Description": e.principal_diagnosis[2] if e.principal_diagnosis else None,
                "Base_Encounter_Cost": round(self.rng.uniform(120, 900), 2),
                "Total_Claim_Cost": None,
                "Payer_Coverage": None,
                "ReasonCode": e.principal_diagnosis[1] if e.principal_diagnosis else None,
                "ReasonDescription": e.chief_complaint,
                # our additions
                "facility_id": e.facility_id,
                "unit_id": e.stays[0][0].unit_id if e.stays else f"{e.facility_id}-ED",
                "encounter_class_code": e.act_code,
                "encounter_status": "finished",
                "patient_class": e.patient_class,
                "mrn": mrn,
                # Which EHR produced the row. The client states facilities run a
                # mix of systems, and the casing / date-format divergence below
                # is keyed off this column -- it is what makes standardisation a
                # real problem rather than a formality.
                "source_system": src,
            }
            erow["EncounterClass"] = _case(erow["EncounterClass"], src)
            erow, dup = self.dx.mutate_row(
                "ehr.encounters", erow, "Id",
                required_fields=("Start", "Patient", "EncounterClass"),
                coded_fields=("EncounterClass", "encounter_class_code", "patient_class"),
                date_pairs=(("Start", "Stop"),))
            erow = self.dx.orphan_key("ehr.encounters", erow, "Id", "Patient")
            encounters.append(erow)
            if dup:
                encounters.append(dup)

            if e.is_inpatient:
                arow = {
                    "subject_id": p["Id"], "hadm_id": e.hadm_id,
                    "admittime": e.admittime, "dischtime": e.dischtime,
                    "deathtime": e.deathtime,
                    "admission_type": e.admission_type,
                    "admit_provider_id": self.rng.choice(self.dims.staff)["staff_id"],
                    "admission_location": e.admission_location,
                    "discharge_location": e.discharge_location,
                    "insurance": e.payer["payer_type"] if e.payer else None,
                    "language": self.rng.choices(["ENGLISH", "SPANISH", "OTHER", "?"],
                                                 weights=[0.83, 0.09, 0.05, 0.03])[0],
                    "marital_status": p["Marital"],
                    "race": p["Race"],
                    "edregtime": e.ed_arrival, "edouttime": e.ed_departure,
                    "hospital_expire_flag": e.hospital_expire_flag,
                    # our additions
                    "facility_id": e.facility_id,
                    "admission_type_code": e.admission_type_code,
                    "admit_decision_time": e.admit_decision_time,
                    "hospital_service": e.hospital_service,
                    "transferred_in_within_6h": e.transferred_in_within_6h,
                    "is_readmission": e.is_readmission,
                    "is_planned_readmission": e.is_planned_readmission,
                    "index_encounter_id": e.index_encounter_id,
                    "encounter_id": e.encounter_id,
                    "drg_code": e.drg_code,
                    "source_system": src,
                }
                arow, dup = self.dx.mutate_row(
                    "ehr.admissions", arow, "hadm_id",
                    required_fields=("admittime", "admission_type"),
                    coded_fields=("admission_type", "admission_location", "discharge_location"),
                    date_pairs=(("admittime", "dischtime"),))
                admissions.append(arow)
                if dup:
                    admissions.append(dup)

                # transfers: ed -> admit -> transfer* -> discharge
                tseq = 0
                if e.ed_arrival and not e.is_outpatient:
                    tseq += 1
                    transfers.append({
                        "subject_id": p["Id"], "hadm_id": e.hadm_id,
                        "transfer_id": f"{e.hadm_id}-{tseq}", "eventtype": "ed",
                        "careunit": "Emergency Department",
                        "intime": e.ed_arrival, "outtime": e.ed_departure,
                        "facility_id": e.facility_id,
                        "unit_id": f"{e.facility_id}-ED",
                    })
                for i, (unit, tin, tout) in enumerate(e.stays):
                    tseq += 1
                    transfers.append({
                        "subject_id": p["Id"], "hadm_id": e.hadm_id,
                        "transfer_id": f"{e.hadm_id}-{tseq}",
                        "eventtype": "admit" if i == 0 else "transfer",
                        "careunit": unit.unit_name,
                        "intime": tin, "outtime": tout,
                        "facility_id": e.facility_id, "unit_id": unit.unit_id,
                    })
                tseq += 1
                transfers.append({
                    "subject_id": p["Id"], "hadm_id": e.hadm_id,
                    "transfer_id": f"{e.hadm_id}-{tseq}", "eventtype": "discharge",
                    "careunit": None, "intime": e.dischtime, "outtime": None,
                    "facility_id": e.facility_id,
                    "unit_id": e.stays[-1][0].unit_id if e.stays else None,
                })

            # An outpatient appointment also records an arrival, but it is not
            # an ED stay and must not land in the ED feed -- it would corrupt
            # every ED wait-time measure with clinic visits.
            if e.ed_arrival and not e.is_outpatient:
                tv = e.triage_vitals
                srow = {
                    "subject_id": p["Id"], "stay_id": e.encounter_id,
                    # MIMIC-IV-ED has no encounter concept, so edstays carries no
                    # encounter_id and hadm_id is null for anyone not admitted --
                    # which left ~74% of ED stays with no path to the encounter
                    # they belong to. This column is ours, and it is what makes a
                    # single visit fact possible for ED-discharged patients.
                    "encounter_id": e.encounter_id,
                    "hadm_id": e.hadm_id,
                    "intime": e.ed_arrival, "outtime": e.ed_departure,
                    "gender": "M" if p["Gender"] == "male" else ("F" if p["Gender"] == "female" else "U"),
                    "race": p["Race"],
                    "arrival_transport": e.arrival_transport,
                    "disposition": e.ed_disposition,
                    # MIMIC-IV-ED triage columns
                    "temperature": tv.get("temperature"), "heartrate": tv.get("heartrate"),
                    "resprate": tv.get("resprate"), "o2sat": tv.get("o2sat"),
                    "sbp": tv.get("sbp"), "dbp": tv.get("dbp"),
                    "pain": tv.get("pain"), "acuity": tv.get("acuity"),
                    "chiefcomplaint": e.chief_complaint,
                    # our additions -- MIMIC-IV-ED has NEITHER of these, and
                    # OP-18 / ED-1 / ED-2 cannot be computed without them
                    "triage_time": e.triage_time,
                    # first physician/APP contact -- door-to-doctor. Null when the
                    # patient left without being seen, which is the correct value
                    # and the case the measure must exclude.
                    "provider_seen_time": e.provider_seen_time,
                    "admit_decision_time": e.admit_decision_time,
                    "facility_id": e.facility_id,
                    "source_system": src,
                }
                srow["disposition"] = _case(srow["disposition"], src)
                srow, dup = self.dx.mutate_row(
                    "ehr.ed_stays", srow, "stay_id",
                    required_fields=("intime", "acuity", "disposition"),
                    coded_fields=("disposition", "arrival_transport"),
                    numeric_fields=("heartrate", "sbp", "o2sat"),
                    date_pairs=(("intime", "outtime"),))
                ed_stays.append(srow)
                if dup:
                    ed_stays.append(dup)

            # Scheduled outpatient and same-day-surgery visits. Separate table
            # because the measure is different: the wait clock starts at the
            # APPOINTMENT time, not at arrival, so a patient who turns up
            # twenty minutes early is not waiting twenty minutes longer.
            if e.is_outpatient:
                orow = {
                    "visit_id": e.encounter_id,
                    "encounter_id": e.encounter_id,
                    "subject_id": p["Id"],
                    "facility_id": e.facility_id,
                    "unit_id": e.outpatient_unit.unit_id if e.outpatient_unit else None,
                    "clinic_type": e.outpatient_unit.unit_code if e.outpatient_unit else None,
                    "appointment_time": e.appointment_time,
                    "arrival_time": e.ed_arrival,
                    "provider_seen_time": e.provider_seen_time,
                    "departure_time": e.ed_departure,
                    "seen_by_provider_id": self.rng.choice(self.dims.staff)["staff_id"],
                    "visit_status": "NO SHOW" if e.is_no_show else (
                        "ADMITTED" if e.is_inpatient else "COMPLETED"),
                    "is_no_show": e.is_no_show,
                    "escalated_to_inpatient": bool(e.is_inpatient),
                    "primary_diagnosis_code": e.diagnoses[0][1] if e.diagnoses else None,
                    "payer_id": e.payer["payer_id"] if e.payer else None,
                    "mrn": mrn,
                    "source_system": src,
                }
                orow["visit_status"] = _case(orow["visit_status"], src)
                orow, dup = self.dx.mutate_row(
                    "ehr.outpatient_visits", orow, "visit_id",
                    required_fields=("appointment_time", "clinic_type"),
                    coded_fields=("visit_status", "clinic_type"),
                    date_pairs=(("appointment_time", "departure_time"),))
                outpatient.append(orow)
                if dup:
                    outpatient.append(dup)

            for seq, code, title, cohort in e.diagnoses:
                diagnoses.append({
                    "subject_id": p["Id"], "hadm_id": e.hadm_id or e.encounter_id,
                    "encounter_id": e.encounter_id,
                    "seq_num": seq, "icd_code": code, "icd_version": 10,
                    "icd_title": title, "hrrp_cohort": cohort,
                    "facility_id": e.facility_id,
                })

        self._write(f"ehr/outpatient_visits/ehr_outpatient_visits_{d}.csv", outpatient)
        self.stats["ehr_outpatient_visits"] += len(outpatient)
        self._write(f"ehr/patients/ehr_patients_{d}.csv", patients)
        self._write(f"ehr/encounters/ehr_encounters_{d}.csv", encounters)
        self._write(f"ehr/admissions/ehr_admissions_{d}.csv", admissions)
        self._write(f"ehr/transfers/ehr_transfers_{d}.csv", transfers)
        self._write(f"ehr/ed_stays/ehr_ed_stays_{d}.csv", ed_stays)
        self._write(f"ehr/diagnoses/ehr_diagnoses_{d}.csv", diagnoses)
        self.stats["ehr_encounters"] += len(encounters)
        self.stats["ehr_admissions"] += len(admissions)
        self.stats["ehr_ed_stays"] += len(ed_stays)

    # =====================================================================
    # Source 2 — Billing & claims (daily CSV, cloud files)
    # =====================================================================
    def emit_claims(self, run_date):
        d = _dstr(run_date)
        headers, lines, remits, adjustments = [], [], [], []

        # claims submitted today = discharges CODING_LAG_DAYS ago
        for lag in range(C.CODING_LAG_DAYS[0], C.CODING_LAG_DAYS[1] + 1):
            disch_day = run_date - timedelta(days=lag)
            # don't bill an encounter the EHR feed never emitted (see emit_from)
            if self.emit_from and disch_day < self.emit_from:
                continue
            for e in self.sim.by_close_date.get(disch_day, []):
                if not (e.dischtime or e.ed_departure):
                    continue
                if e.ed_disposition == "LEFT WITHOUT BEING SEEN":
                    continue
                if not e.payer:
                    continue
                # deterministic lag choice so a claim is emitted exactly once
                if (hash(e.encounter_id) % (C.CODING_LAG_DAYS[1] - C.CODING_LAG_DAYS[0] + 1)) \
                        != (lag - C.CODING_LAG_DAYS[0]):
                    continue
                self._build_claim(e, run_date, headers, lines, remits, adjustments)

        self._write(f"claims/claim_header/claim_header_{d}.csv", headers)
        self._write(f"claims/claim_line/claim_line_{d}.csv", lines)
        self._write(f"claims/remit/remit_{d}.csv", remits)
        self._write(f"claims/remit_adjustment/remit_adjustment_{d}.csv", adjustments)
        self.stats["claims"] += len(headers)
        self.stats["denials"] += sum(1 for r in remits if r.get("claim_status_code") == "4")

    def _build_claim(self, e, submission_date, headers, lines, remits, adjustments):
        pcn = f"PCN{e.encounter_id[3:]}"
        inpatient = e.is_inpatient
        # Type of Bill: 011x inpatient, 013x outpatient
        tob = "0111" if inpatient else "0131"
        dx_principal = e.principal_diagnosis

        # line items
        claim_lines, total = [], 0.0
        if inpatient:
            nights = max(1, int(math.ceil(e.los_days or 1)))
            unit = e.stays[0][0] if e.stays else None
            rev = "0200" if (unit and unit.is_critical_care) else "0121"
            rate = 3850.0 if rev == "0200" else 1420.0
            claim_lines.append((rev, None, nights, round(rate * nights, 2)))
            for _ in range(self.rng.randint(2, 7)):
                rev = self.rng.choice(["0250", "0258", "0300", "0320", "0360", "0410", "0730"])
                hcpcs = self.rng.choice(list(R.HCPCS_SEED)) if rev.startswith("025") else None
                units = self.rng.randint(1, 12)
                claim_lines.append((rev, hcpcs, units, round(self.rng.uniform(45, 2400), 2)))
        else:
            claim_lines.append(("0450", None, 1, round(self.rng.uniform(480, 2900), 2)))
            for _ in range(self.rng.randint(1, 4)):
                rev = self.rng.choice(["0250", "0300", "0320", "0730", "0460"])
                claim_lines.append((rev, self.rng.choice(list(R.HCPCS_SEED)) if rev == "0250" else None,
                                    self.rng.randint(1, 4), round(self.rng.uniform(35, 780), 2)))
        total = round(sum(l[3] for l in claim_lines), 2)

        status_code = R.DISCHARGE_TO_STATUS_CODE.get(e.discharge_location, "01") if inpatient else "01"
        adm_src = "7" if e.ed_arrival else self.rng.choice(["1", "2", "4", "D", "E"])

        hrow = {
            "patient_control_number": pcn,
            "encounter_id": e.encounter_id,
            "hadm_id": e.hadm_id,
            "subject_id": e.patient["Id"],
            "facility_id": e.facility_id,
            "total_charge_amount": total,
            "claim_filing_indicator_code": e.payer["claim_filing_indicator_code"],
            "payer_id": e.payer["payer_id"],
            # The clearinghouse passes the payer name through as it received it,
            # so one payer arrives spelled several ways and Silver has to fold
            # them to a canonical name. payer_id is always clean -- that is the
            # key; the name is the standardisation problem.
            "payer_name": self.rng.choice(
                C.PAYER_NAME_VARIANTS.get(e.payer["payer_id"], [e.payer["payer_name"]])),
            "type_of_bill": tob,
            "statement_date_from": (e.admittime or e.ed_arrival),
            "statement_date_to": (e.dischtime or e.ed_departure),
            "admission_date_and_hour": e.admittime,
            "discharge_time": e.dischtime,
            "admission_type_code": e.admission_type_code,
            "admission_source_code": adm_src,
            "patient_status_code": status_code,
            # Family lookup plus a severity draw, NOT the CMS grouper -- see
            # refdata.MSDRG_FAMILIES. Outpatient claims carry no DRG.
            "drg_code": e.drg_code,
            "principal_diagnosis": dx_principal[1] if dx_principal else None,
            "admitting_diagnosis": dx_principal[1] if dx_principal else None,
            "other_diagnoses": "|".join(d[1] for d in e.diagnoses[1:8]),
            "attending_provider_npi": (self.rng.choice(
                [s["npi"] for s in self.dims.staff if s["npi"]]) or None),
            "medical_record_number": self.dims.mrn_for(e.patient, e.facility_id),
            "prior_authorization_number": (f"AUTH{self.rng.randint(10**7, 10**8-1)}"
                                           if self.rng.random() < 0.42 else None),
            # HRRP penalty exposure. True when this claim is for a readmission
            # inside the 30-day window of an eligible index admission.
            "is_readmission_related": bool(e.is_readmission and not e.is_planned_readmission),
            "submission_date": submission_date,
        }
        hrow, dup = self.dx.mutate_row(
            "claims.claim_header", hrow, "patient_control_number",
            required_fields=("total_charge_amount", "payer_id", "principal_diagnosis"),
            coded_fields=("type_of_bill", "patient_status_code", "admission_source_code"),
            numeric_fields=("total_charge_amount",),
            date_pairs=(("statement_date_from", "statement_date_to"),))
        hrow = self.dx.orphan_key("claims.claim_header", hrow,
                                  "patient_control_number", "encounter_id")
        headers.append(hrow)
        if dup:
            headers.append(dup)

        for i, (rev, hcpcs, units, charge) in enumerate(claim_lines, start=1):
            lrow = {
                "patient_control_number": pcn,
                "line_control_number": f"{pcn}-{i:03d}",
                "revenue_code": rev,
                "revenue_code_description": R.REVENUE_CODES.get(rev),
                "procedure_code": hcpcs,
                "procedure_code_qualifier": "HC" if hcpcs else None,
                "procedure_description": R.HCPCS_SEED.get(hcpcs) if hcpcs else None,
                "line_charge_amount": charge,
                "unit_type": "DA" if rev in ("0111", "0121", "0200", "0210") else "UN",
                "unit_count": units,
                "non_covered_amount": 0.0,
                "service_date_from": (e.admittime or e.ed_arrival),
                "service_date_to": (e.dischtime or e.ed_departure),
            }
            lrow, ldup = self.dx.mutate_row(
                "claims.claim_line", lrow, "line_control_number",
                required_fields=("revenue_code", "line_charge_amount"),
                coded_fields=("revenue_code",),
                numeric_fields=("line_charge_amount", "unit_count"))
            lines.append(lrow)
            if ldup:
                lines.append(ldup)

        # ---- adjudication ------------------------------------------------
        ptype = e.payer["payer_type"]
        if ptype == "Self-Pay":
            return
        denial_p = min(0.95, C.BLENDED_INITIAL_DENIAL_RATE
                       * C.DENIAL_MULTIPLIER_BY_PAYER_TYPE.get(ptype, 1.0))
        denied = self.rng.random() < denial_p

        med, sigma = C.DAYS_TO_PAYMENT_BY_PAYER_TYPE[ptype]
        days = max(e.payer["prompt_pay_days"] * 0.6,
                   med * math.exp(self.rng.gauss(0, sigma)))
        remit_date = submission_date + timedelta(days=int(days))

        contractual = round(total * self.rng.uniform(0.28, 0.62), 2)
        if denied:
            paid, patient_resp = 0.0, 0.0
            status = "4"
        else:
            patient_resp = round((total - contractual) * self.rng.uniform(0.0, 0.22), 2)
            paid = round(max(0.0, total - contractual - patient_resp), 2)
            status = "1"

        rrow = {
            "patient_control_number": pcn,
            "payer_id": e.payer["payer_id"],
            "claim_status_code": status,
            "claim_status_description": R.CLAIM_STATUS_CODE[status],
            "total_claim_charge_amount": total,
            "claim_payment_amount": paid,
            "patient_responsibility_amount": patient_resp,
            "payer_claim_control_number": f"ICN{self.rng.randint(10**11, 10**12-1)}",
            "drg_code": e.drg_code,
            "drg_weight": e.drg_weight,
            "check_eft_trace_number": f"EFT{self.rng.randint(10**8, 10**9-1)}",
            "payment_method_code": self.rng.choice(["ACH", "CHK"]),
            "check_date": remit_date,
            "remit_date": remit_date,
            "is_appealed": None,
            "is_overturned_on_appeal": None,
        }
        if denied:
            rrow["is_appealed"] = self.rng.random() < 0.62
            if rrow["is_appealed"]:
                rrow["is_overturned_on_appeal"] = self.rng.random() < C.APPEAL_OVERTURN_RATE
        rrow, rdup = self.dx.mutate_row(
            "claims.remit", rrow, "patient_control_number",
            required_fields=("claim_status_code", "claim_payment_amount"),
            coded_fields=("claim_status_code",),
            numeric_fields=("claim_payment_amount", "patient_responsibility_amount"))
        remits.append(rrow)
        if rdup:
            remits.append(rdup)

        adj_seq = 0
        def add_adj(group, code, desc, amount, remark=None):
            nonlocal adj_seq
            adj_seq += 1
            adjustments.append({
                "patient_control_number": pcn,
                "adjustment_seq": adj_seq,
                "group_code": group,
                "group_code_description": R.ADJUSTMENT_GROUP_CODE[group],
                "reason_code": code,
                "reason_code_description": desc,
                "amount": amount,
                "quantity": None,
                "remark_code": remark,
                "remark_code_description": R.RARC.get(remark),
                "is_denial": group in ("CO", "OA") and code in
                             [c for c, _g, _d, _w in R.DENIAL_CARCS],
            })

        # contractual write-off is present on paid AND denied claims -- and is
        # NOT a denial. Keeping these distinct is what stops the Gold layer
        # overstating the denial rate.
        add_adj("CO", "45", next(d for c, g, d, isd, w in R.CARC if c == "45"), contractual)
        if denied:
            code, group, desc, _w = self.rng.choices(
                R.DENIAL_CARCS, weights=[w for _c, _g, _d, w in R.DENIAL_CARCS])[0]
            remark = "MA130" if code == "16" else None   # CARC 16 requires a RARC
            add_adj(group, code, desc, round(total - contractual, 2), remark)
        else:
            if patient_resp > 0:
                pr = self.rng.choice([("1", "Deductible Amount"),
                                      ("2", "Coinsurance Amount"),
                                      ("3", "Co-payment Amount")])
                add_adj("PR", pr[0], pr[1], patient_resp)

    # =====================================================================
    # Source 3 — Pharmacy inventory (daily snapshot, cloud files)
    # =====================================================================
    def _init_inventory(self, run_date):
        self._inventory = {}
        for f in self.dims.facilities:
            locations = [f"{f.facility_id}-PHARM-MAIN"]
            for u in self.dims.units_by_facility.get(f.facility_id, []):
                if u.unit_code in ("ED", "MICU", "SICU", "CVICU", "MS", "TELE"):
                    locations.append(f"{f.facility_id}-ADC-{u.unit_code}")
            scale = max(0.25, (f.licensed_beds or 40) / 400.0)
            for loc in locations:
                is_adc = "ADC" in loc
                for drug in self.dims.drugs:
                    if is_adc and self.rng.random() < 0.55:
                        continue     # cabinets carry a subset
                    daily = max(0.4, drug["usage_weight"] * scale * (0.25 if is_adc else 1.0))
                    doh = self.rng.uniform(*C.DAYS_ON_HAND_TARGET)
                    par = max(4, int(daily * doh))
                    self._inventory[(loc, drug["ndc11"])] = {
                        "facility_id": f.facility_id,
                        "location_id": loc,
                        "drug": drug,
                        "avg_daily_usage": daily,
                        "par_level": par,
                        "reorder_point": max(2, int(daily * C.REORDER_POINT_DAYS)),
                        "safety_stock": max(1, int(daily * C.SAFETY_STOCK_DAYS)),
                        "qty_on_hand": int(par * self.rng.uniform(0.55, 1.0)),
                        "qty_on_order": 0,
                        "order_eta": None,
                        "lot_number": f"L{self.rng.randint(100000, 999999)}",
                        "expiration_date": run_date + timedelta(days=self.rng.randint(45, 900)),
                    }

    def _tick_shortages(self, run_date):
        for drug in self.dims.drugs:
            ndc = drug["ndc11"]
            st = self._shortages.get(ndc)
            if st and st["until"] <= run_date:
                self._shortages.pop(ndc)
            elif not st:
                p = C.SHORTAGE_PROBABILITY_PER_DRUG / 90.0
                if drug["is_shortage_prone"]:
                    p *= 4.0
                if self.rng.random() < p:
                    self._shortages[ndc] = {
                        "until": run_date + timedelta(days=self.rng.randint(*C.SHORTAGE_DURATION_DAYS)),
                        "reason": self.rng.choice(C.SHORTAGE_REASONS),
                    }

    def emit_inventory(self, run_date):
        if self._inventory is None:
            self._init_inventory(run_date)
        self._tick_shortages(run_date)
        d = _dstr(run_date)
        rows = []

        for (loc, ndc), inv in self._inventory.items():
            drug = inv["drug"]
            shortage = self._shortages.get(ndc)

            # consumption
            used = max(0, int(self.rng.gauss(inv["avg_daily_usage"],
                                             inv["avg_daily_usage"] * 0.35)))
            inv["qty_on_hand"] = max(0, inv["qty_on_hand"] - used)

            # receipts
            if inv["order_eta"] and inv["order_eta"] <= run_date:
                inv["qty_on_hand"] += inv["qty_on_order"]
                inv["qty_on_order"] = 0
                inv["order_eta"] = None
                inv["lot_number"] = f"L{self.rng.randint(100000, 999999)}"
                inv["expiration_date"] = run_date + timedelta(days=self.rng.randint(120, 900))

            # reorder
            if inv["qty_on_hand"] <= inv["reorder_point"] and inv["qty_on_order"] == 0:
                inv["qty_on_order"] = inv["par_level"] - inv["qty_on_hand"]
                lead = self.rng.randint(1, 4)
                if shortage:
                    lead += self.rng.randint(4, 21)      # shortage stretches lead time
                inv["order_eta"] = run_date + timedelta(days=lead)

            doh = (inv["qty_on_hand"] / inv["avg_daily_usage"]) if inv["avg_daily_usage"] else None
            row = {
                "snapshot_date": run_date,
                "counting_datetime": datetime.combine(run_date, time(2, 15)),
                "count_type": "snapshot",
                "facility_id": inv["facility_id"],
                "location_id": loc,
                "ndc11": ndc,
                "product_ndc": drug["product_ndc"],
                "gtin14": drug["gtin14"],
                "rxcui_scd": drug["rxcui_scd"],
                "drug_name": drug["non_proprietary_name"],
                "dosage_form_name": drug["dosage_form_name"],
                "route_name": drug["route_name"],
                "pharm_classes": drug["pharm_classes"],
                "lot_number": inv["lot_number"],
                "expiration_date": inv["expiration_date"],
                "qty_on_hand": inv["qty_on_hand"],
                "base_unit": "EA",
                "qty_on_order": inv["qty_on_order"],
                "par_level": inv["par_level"],
                "reorder_point": inv["reorder_point"],
                "safety_stock": inv["safety_stock"],
                "avg_daily_usage_30d": round(inv["avg_daily_usage"], 2),
                "days_on_hand": round(doh, 2) if doh is not None else None,
                "abc_class": drug["abc_class"],
                "is_controlled": drug["is_controlled"],
                "dea_schedule": drug["dea_schedule"],
                "is_high_alert": drug["is_high_alert"],
                "shortage_status": "Currently in Shortage" if shortage else "Available",
                "shortage_reason": shortage["reason"] if shortage else None,
                "unit_cost": drug["unit_cost"],
                "extended_value": round(inv["qty_on_hand"] * drug["unit_cost"], 2),
                "last_count_variance": self.rng.choice([0, 0, 0, 0, -1, 1, -2, 2]),
                # When this line was last replenished. Needed to tell a genuine
                # stockout from an item simply awaiting its next delivery.
                "last_restocked_at": (run_date - timedelta(days=self.rng.randint(0, 21))),
                "is_stockout": inv["qty_on_hand"] == 0,
            }
            row, dup = self.dx.mutate_row(
                "pharmacy.inventory", row, "ndc11",
                required_fields=("qty_on_hand", "lot_number", "expiration_date"),
                coded_fields=("shortage_status", "abc_class", "dea_schedule"),
                numeric_fields=("qty_on_hand", "par_level", "unit_cost"))
            rows.append(row)
            if dup:
                rows.append(dup)

        self._write(f"pharmacy/inventory/pharmacy_inventory_{d}.csv", rows)
        self.stats["inventory_rows"] += len(rows)
        self.stats["stockouts"] += sum(1 for r in rows if r.get("is_stockout"))

    # =====================================================================
    # Source 4 — Bed capacity snapshots (hourly + NHSN-shaped weekly roll-up)
    # =====================================================================
    def emit_bed_capacity(self, run_date):
        day = run_date - timedelta(days=1)
        d = _dstr(run_date)
        rows = [dict(s) for s in self.sim.snapshots_by_date.get(day, [])]
        if not rows:
            return
        out = []
        for r in rows:
            r, dup = self.dx.mutate_row(
                "beds.hourly_snapshot", r, "unit_id",
                required_fields=("occupied_beds", "staffed_beds"),
                numeric_fields=("occupied_beds", "staffed_beds", "available_beds"))
            out.append(r)
            if dup:
                out.append(dup)
        self._write(f"beds/hourly_snapshot/bed_snapshot_hourly_{d}.csv", out)
        self.stats["bed_snapshots"] += len(out)

    def emit_nhsn_weekly(self, week_ending: date):
        """
        NHSN Hospital Respiratory Data shape: facility x week, capacity and
        occupancy measured as of the WEDNESDAY of the reporting week.

        This roll-up must reconcile against the hourly detail. That
        reconciliation is a free, genuinely meaningful DQ check -- and the kind
        of thing the client will actually poke at.
        """
        wednesday = week_ending - timedelta(days=(week_ending.weekday() - 2) % 7)
        # The measurement Wednesday can fall before the extract window opens --
        # a Sunday early in the window looks back to the previous Wednesday. A
        # roll-up measured on a day we never delivered hourly detail for cannot
        # be reconciled against anything, so it is not a roll-up worth shipping:
        # it would read as a permanent reconciliation failure rather than a
        # genuine variance.
        if self.emit_from and wednesday < self.emit_from:
            return
        rows = []
        for f in self.dims.facilities:
            units = self.dims.inpatient_units(f.facility_id)
            if not units:
                continue
            snaps = [s for s in self.sim.snapshots_by_date.get(wednesday, [])
                     if s["facility_id"] == f.facility_id]
            if not snaps:
                continue
            icu_codes = {"MICU", "SICU", "CVICU", "NICU"}
            peds_codes = {"PEDS", "NICU"}
            def agg(pred, field):
                return sum(s[field] for s in snaps if pred(s))
            rows.append({
                "nhsn_org_id": f.facility_id,
                "facility_name": f.name,
                "week_ending_date": week_ending,
                "collection_date": wednesday,
                "all_hospital_inpatient_beds": agg(lambda s: True, "staffed_beds") // 24,
                "all_hospital_inpatient_occupancy": agg(lambda s: True, "occupied_beds") // 24,
                "all_adult_inpatient_beds": agg(lambda s: s["unit_code"] not in peds_codes, "staffed_beds") // 24,
                "all_adult_inpatient_occupancy": agg(lambda s: s["unit_code"] not in peds_codes, "occupied_beds") // 24,
                "all_pediatric_inpatient_beds": agg(lambda s: s["unit_code"] in peds_codes, "staffed_beds") // 24,
                "all_pediatric_inpatient_occupancy": agg(lambda s: s["unit_code"] in peds_codes, "occupied_beds") // 24,
                "all_icu_beds": agg(lambda s: s["unit_code"] in icu_codes, "staffed_beds") // 24,
                "all_icu_bed_occupancy": agg(lambda s: s["unit_code"] in icu_codes, "occupied_beds") // 24,
                "adult_icu_beds": agg(lambda s: s["unit_code"] in (icu_codes - peds_codes), "staffed_beds") // 24,
                "adult_icu_bed_occupancy": agg(lambda s: s["unit_code"] in (icu_codes - peds_codes), "occupied_beds") // 24,
                "pediatric_icu_beds": agg(lambda s: s["unit_code"] == "NICU", "staffed_beds") // 24,
                "pediatric_icu_bed_occupancy": agg(lambda s: s["unit_code"] == "NICU", "occupied_beds") // 24,
            })
        if rows:
            self.batch.write_csv(
                f"beds/nhsn_weekly/bed_capacity_nhsn_weekly_{_dstr(week_ending)}.csv", rows)
            self.stats["nhsn_weeks"] += 1

    # =====================================================================
    # Source 5 — Staff schedules (.xlsx, SharePoint document library)
    # =====================================================================
    def emit_staff_schedule(self, week_start: date):
        """
        A human-maintained operational spreadsheet, with the mess that implies:
        a title block above the real header, facility name in a cell rather than
        a column, mixed date formats, a free-text Notes column, trailing blanks.
        That mess is the point -- it exercises a different ingestion path from
        the machine-generated CSV drops.
        """
        from openpyxl import Workbook

        # The roster covers the week ahead and is overwritten daily as actuals
        # come in, so the copy that lands in the document library is a mid-week
        # snapshot: shifts before the cutoff have actual hours, shifts after it
        # are still only scheduled. Those trailing nulls are legitimate, and
        # they are the material a DQ gate needs to prove it can tell a real
        # null from a missing value.
        self._roster_cutoff = datetime.combine(week_start + timedelta(days=5), time(6, 0))

        for f in self.dims.facilities:
            units = self.dims.units_by_facility.get(f.facility_id, [])
            if not units:
                continue
            wb = Workbook()
            ws = wb.active
            ws.title = "Roster"

            ws["A1"] = f"{f.name} — Nursing Staff Schedule"
            ws["A2"] = f"Week commencing {week_start.strftime('%d %B %Y')}"
            ws["A3"] = "Prepared by: Nurse Staffing Office     CONFIDENTIAL"
            ws["A4"] = ""

            header = ["Facility ID", "Unit", "Unit Code", "Work Date", "Shift",
                      "Shift Start", "Shift End", "Staff ID", "Name", "Job Code",
                      "Employment Type", "Scheduled Hours", "Actual Hours",
                      "Status", "Overtime", "Called Out", "Floated In",
                      "Census", "Notes"]
            ws.append(header)

            rows_written = 0
            for offset in range(7):
                day = week_start + timedelta(days=offset)
                is_weekend = day.weekday() >= 5
                for u in units:
                    census = self.sim.census(day, u.unit_id) or 0
                    if u.unit_code == "ED":
                        census = max(4, int(f.ed_arrivals_per_day / 12))
                    if census == 0 and u.unit_code != "ED":
                        continue
                    for shift_code, start_h, length in C.SHIFTS:
                        required = max(1, math.ceil(census / u.nurse_patient_ratio_target))
                        bias = C.UNDERSTAFF_BIAS.get(
                            (f.facility_type, u.unit_code, shift_code), C.UNDERSTAFF_DEFAULT)
                        if is_weekend:
                            bias *= C.WEEKEND_STAFF_PENALTY
                        if shift_code == "OC":
                            # On-call is a fixed cover team, not census-scaled
                            # presence. Rostering it against census produces
                            # nonsense ratios -- one on-call nurse "covering" a
                            # 50-patient unit -- and a staffing-adequacy KPI
                            # must not be computed against it at all.
                            scheduled = self.rng.choice([1, 1, 2])
                        else:
                            # D/E/N each cover their own 8h block, so each is
                            # staffed to the census ratio; the per-shift share
                            # only encodes that nights run slightly leaner.
                            cover = C.SHIFT_COVERAGE_SHARE.get(shift_code, 1.0)
                            scheduled = max(1, int(round(required * bias * cover)))
                        pool = self.dims.nurses_for_unit(f.facility_id, u.unit_id)
                        if not pool:
                            continue
                        picks = [self.rng.choice(pool) for _ in range(scheduled)]
                        shift_start = datetime.combine(day, time(start_h, 0))
                        shift_end = shift_start + timedelta(hours=length)
                        for s in picks:
                            called_out = self.rng.random() < C.CALL_OUT_RATE
                            overtime = (not called_out) and self.rng.random() < C.OVERTIME_RATE
                            actual = 0.0 if called_out else (length + (self.rng.choice([2, 4]) if overtime else 0))
                            # A shift that has not finished yet has no actual
                            # hours. This is a LEGITIMATE null, distinct from
                            # the nulls the defect injector adds, and it is the
                            # material a DQ gate needs in order to prove it does
                            # not false-positive on one.
                            not_yet_worked = shift_end > self._roster_cutoff
                            status = ("scheduled" if not_yet_worked else
                                      ("absent" if called_out else "completed"))
                            if not_yet_worked and self.rng.random() < 0.06:
                                status = self.rng.choice(["swapped", "cancelled"])
                            # mixed date formats, as a real spreadsheet has
                            if self.rng.random() < 0.08:
                                work_date = day.strftime("%m/%d/%Y")
                            else:
                                work_date = day.strftime("%Y-%m-%d")
                            ws.append([
                                f.facility_id, u.unit_name, u.unit_code, work_date,
                                shift_code,
                                shift_start.strftime("%H:%M"), shift_end.strftime("%H:%M"),
                                s["staff_id"], f"{s['last_name']}, {s['first_name']}",
                                s["job_code"], s["employment_type"],
                                float(length),
                                None if not_yet_worked else float(actual),
                                status,
                                "Y" if overtime else "", "Y" if called_out else "",
                                "Y" if self.rng.random() < 0.05 else "",
                                census,
                                self.rng.choice(["", "", "", "", "orientee paired",
                                                 "agency", "float pool", "double shift"]),
                            ])
                            rows_written += 1
            # trailing blank rows, as a real spreadsheet has
            for _ in range(self.rng.randint(1, 4)):
                ws.append([])

            buf = io.BytesIO()
            wb.save(buf)
            fname = (f"staff_schedule_{f.facility_id}_"
                     f"{week_start.strftime('%Y%m%d')}.xlsx")
            self.batch.write_bytes(f"sharepoint/staff_schedules/{fname}", buf.getvalue())
            self.stats["staff_schedule_rows"] += rows_written
            self.stats["staff_schedule_files"] += 1

    # =====================================================================
    # Source 6 — Patient vitals stream (Kafka)
    # =====================================================================
    def emit_vitals(self, run_date):
        """
        One event per parameter reading (tall/EAV, like MIMIC chartevents)
        rather than a wide row: handles missing parameters naturally, matches
        how monitors actually emit, and gives Silver real pivoting work.

        Cadence follows eICU: interfaced as 1-minute averages, archived as
        5-minute medians. We emit the 5-minute archive.
        """
        day = run_date - timedelta(days=1)
        monitored = []
        for e in self.sim.active_by_date.get(day, []):
            if not e.stays:
                continue
            for unit, tin, tout in e.stays:
                if not unit.is_monitored or tin is None:
                    continue
                s = max(tin, datetime.combine(day, time(0, 0)))
                x = min(tout or datetime.combine(day + timedelta(days=1), time(0, 0)),
                        datetime.combine(day + timedelta(days=1), time(0, 0)))
                if x > s:
                    monitored.append((e, unit, s, x))

        for e, unit, s, x in monitored:
            severity = (6 - (e.esi_acuity or 3)) / 5.0
            deteriorating = self.rng.random() < (0.06 + 0.22 * severity)
            bed_id = f"{unit.unit_id}-B{self.rng.randint(1, max(1, unit.staffed_beds)):02d}"
            device_id = f"MON-{unit.unit_code}-{self.rng.randint(1, 40):03d}"
            t = s.replace(second=0, microsecond=0)
            step = timedelta(minutes=C.VITALS_INTERVAL_MIN)
            n_steps = 0
            while t < x:
                n_steps += 1
                # signal dropout, as real monitors have
                if self.rng.random() < 0.012:
                    t += step
                    continue
                frac = (t - s).total_seconds() / max(1.0, (x - s).total_seconds())
                drift = (frac * 1.6) if deteriorating else 0.0
                for loinc, name, uom, mean, sd, lo, hi, dec in C.VITAL_PARAMS:
                    shift = self._vital_shift(loinc, severity + drift)
                    val = mean + shift + self.rng.gauss(0, sd)
                    val = max(lo, min(hi, val))
                    val = round(val, dec) if dec else int(round(val))
                    is_artifact = self.rng.random() < 0.008
                    if is_artifact:
                        val = round(val * self.rng.choice([0.35, 2.4]), dec) if dec else int(val * self.rng.choice([0.35, 2.4]))
                    ev = {
                        "event_id": str(uuid.uuid5(
                            uuid.NAMESPACE_OID,
                            f"vitals|{e.encounter_id}|{device_id}|{loinc}|{t.isoformat()}")),
                        "event_type": "vitals.reading",
                        "event_time": t,
                        "ingest_time": t + timedelta(seconds=self.rng.randint(1, 9)),
                        "source_system": "PHILIPS_IX_MONITOR",
                        "facility_id": e.facility_id,
                        "schema_version": "1.0",
                        "payload": {
                            "patient_id": e.patient["Id"],
                            "encounter_id": e.encounter_id,
                            "hadm_id": e.hadm_id,
                            "unit_id": unit.unit_id,
                            "bed_id": bed_id,
                            "device_id": device_id,
                            "charttime": t,
                            "loinc_code": loinc,
                            "parameter_name": name,
                            "value_num": val,
                            "value_uom": uom,
                            # A monitor alarm: the reading is outside the band
                            # that scores 3 on NEWS2 for this parameter. This
                            # previously duplicated is_artifact, which made the
                            # column useless -- an artifact is a bad reading, an
                            # alarm is a sick patient, and the operational
                            # dashboard needs the second one.
                            "warning": 1 if self._is_alarm(loinc, val) else 0,
                            "is_artifact": is_artifact,
                        },
                    }
                    for out in self.dx.mutate_event("stream.vitals", ev):
                        self.stream.send("patient-vitals", e.patient["Id"], out)
                        self.stats["vitals_events"] += 1
                t += step

    @staticmethod
    def _vital_shift(loinc, severity):
        """Physiological derangement scaled by severity. [ASSUMPTION]"""
        return {
            "8867-4":  34 * severity,     # tachycardia
            "8480-6": -24 * severity,     # hypotension
            "8462-4": -11 * severity,
            "9279-1":  10 * severity,     # tachypnoea
            "8310-5":  1.1 * severity,    # fever
            "2708-6":  -8 * severity,     # desaturation
        }.get(loinc, 0.0)

    # =====================================================================
    # Source 7 — Prescription issuance stream (Kafka)
    # =====================================================================
    def emit_prescriptions(self, run_date):
        """
        MIMIC-IV hosp.pharmacy + hosp.prescriptions field names, with FHIR
        MedicationRequest supplying the status vocabulary.

        NOTE the deliberate VARCHAR-for-numeric on dose_val_rx / fill_quantity.
        That is authentic MIMIC/EHR messiness and it gives Silver a real
        type-conformance job. Do not "fix" it in the generator.
        """
        day = run_date - timedelta(days=1)
        for e in self.sim.active_by_date.get(day, []):
            if not e.is_inpatient or not e.admittime:
                continue
            s = max(e.admittime, datetime.combine(day, time(0, 0)))
            x = min(e.dischtime or datetime.combine(day + timedelta(days=1), time(0, 0)),
                    datetime.combine(day + timedelta(days=1), time(0, 0)))
            if x <= s:
                continue
            severity = (6 - (e.esi_acuity or 3)) / 5.0
            n_orders = max(1, int(self.rng.gauss(4 + 6 * severity, 2.2)))
            for _ in range(n_orders):
                drug = self.dims.pick_drug()
                entertime = s + timedelta(
                    seconds=self.rng.randint(0, max(1, int((x - s).total_seconds()))))
                verifiedtime = entertime + timedelta(minutes=self.rng.randint(2, 55))
                starttime = verifiedtime + timedelta(minutes=self.rng.randint(5, 180))
                stoptime = starttime + timedelta(hours=self.rng.choice([6, 12, 24, 48, 72]))
                pharmacy_id = self.rng.randint(10**6, 10**7 - 1)
                dose = self.rng.choice([0.5, 1, 2, 4, 5, 10, 12.5, 20, 25, 40, 50, 100, 500, 1000])
                status = self.rng.choices(["active", "inactive", "discontinued"],
                                          weights=[0.72, 0.19, 0.09])[0]
                ev = {
                    "event_id": str(uuid.uuid5(
                        uuid.NAMESPACE_OID,
                        f"rx|{e.encounter_id}|{pharmacy_id}|{drug['ndc11']}|{entertime.isoformat()}")),
                    "event_type": "prescription.issued",
                    "event_time": entertime,
                    "ingest_time": entertime + timedelta(seconds=self.rng.randint(1, 12)),
                    "source_system": "PHARMACY_OMS",
                    "facility_id": e.facility_id,
                    "schema_version": "1.0",
                    "payload": {
                        "subject_id": e.patient["Id"],
                        "hadm_id": e.hadm_id,
                        "encounter_id": e.encounter_id,
                        "pharmacy_id": pharmacy_id,
                        "poe_id": f"{e.hadm_id}-{self.rng.randint(1, 999)}",
                        "poe_seq": self.rng.randint(1, 999),
                        "order_provider_id": self.rng.choice(self.dims.staff)["staff_id"],
                        "drug": drug["non_proprietary_name"],
                        "drug_type": self.rng.choices(["MAIN", "BASE", "ADDITIVE"],
                                                      weights=[0.80, 0.13, 0.07])[0],
                        "formulary_drug_cd": drug["non_proprietary_name"][:12].upper().replace(" ", ""),
                        "gsn": str(self.rng.randint(1000, 99999)),
                        "ndc": drug["ndc11"],
                        "rxcui_scd": drug["rxcui_scd"],
                        "prod_strength": drug["non_proprietary_name"],
                        # VARCHAR on purpose -- see docstring
                        "dose_val_rx": str(dose),
                        "dose_unit_rx": self.rng.choice(["mg", "mL", "UNT", "mEq", "mcg"]),
                        "form_val_disp": str(self.rng.randint(1, 4)),
                        "form_unit_disp": self.rng.choice(["TAB", "VIAL", "BAG", "SYR"]),
                        "doses_per_24_hrs": float(self.rng.choice([1, 2, 3, 4, 6])),
                        "route": drug["route_name"],
                        "frequency": self.rng.choice(["Q24H", "Q12H", "Q8H", "Q6H", "BID", "TID", "ONCE", "PRN"]),
                        "proc_type": self.rng.choice(
                            ["Unit Dose", "IV Piggyback", "Non-formulary", "Large Volume"]),
                        "status": status,
                        "fhir_status": {"active": "active", "inactive": "completed",
                                        "discontinued": "stopped"}[status],
                        "fhir_intent": "order",
                        "fhir_priority": "stat" if severity > 0.7 and self.rng.random() < 0.3 else "routine",
                        "entertime": entertime,
                        "verifiedtime": verifiedtime,
                        "starttime": starttime,
                        "stoptime": stoptime,
                        "dispensation": self.rng.choice(["Main Pharmacy", "Satellite", "ADC"]),
                        "fill_quantity": str(self.rng.randint(1, 6)),
                        "is_controlled": drug["is_controlled"],
                        "dea_schedule": drug["dea_schedule"],
                        "unit_id": e.stays[0][0].unit_id if e.stays else None,
                        "event_subtype": "ordered",
                    },
                }
                for out in self.dx.mutate_event("stream.prescriptions", ev):
                    self.stream.send("prescription-events", e.patient["Id"], out)
                    self.stats["prescription_events"] += 1
